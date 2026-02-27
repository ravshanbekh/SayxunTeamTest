"""
Export service for generating Excel and PDF reports.
Optimized: batch-loads all data in ~4 queries instead of N+1.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import List
from uuid import UUID
from datetime import datetime
import os
import json as _json
import time
import logging

from app.models.result import Result, MCQAnswer, WrittenAnswer
from app.models.user import User
from app.models.test import Test, AnswerKey
from app.config import settings
from app.utils.answer_compare import normalize as _norm, answers_match as _match

logger = logging.getLogger(__name__)


async def _load_export_data(db: AsyncSession, test_id: UUID):
    """
    Batch-load all data needed for export in ~4 queries.
    Returns (test, results, users_dict, mcq_dict, written_dict, written_answers_key)
    """
    t0 = time.time()

    # 1. Test
    stmt = select(Test).where(Test.id == test_id)
    res = await db.execute(stmt)
    test = res.scalars().first()
    t1 = time.time()
    logger.info(f"[EXPORT] Test loaded: {(t1-t0)*1000:.0f}ms")

    # 2. Results
    stmt = select(Result).where(Result.test_id == test_id).order_by(Result.submitted_at)
    res = await db.execute(stmt)
    results = res.scalars().all()
    result_ids = [r.id for r in results]
    user_ids = list(set(r.user_id for r in results))
    t2 = time.time()
    logger.info(f"[EXPORT] {len(results)} results loaded: {(t2-t1)*1000:.0f}ms")

    # 3. All users in one query
    stmt = select(User).where(User.id.in_(user_ids))
    res = await db.execute(stmt)
    users_list = res.scalars().all()
    users_dict = {u.id: u for u in users_list}
    t3 = time.time()
    logger.info(f"[EXPORT] {len(users_list)} users loaded: {(t3-t2)*1000:.0f}ms")

    # 4. All MCQ answers in one query
    stmt = select(MCQAnswer).where(MCQAnswer.result_id.in_(result_ids)).order_by(MCQAnswer.question_number)
    res = await db.execute(stmt)
    all_mcq = res.scalars().all()
    mcq_by_result = {}
    for ans in all_mcq:
        mcq_by_result.setdefault(ans.result_id, []).append(ans)
    t4 = time.time()
    logger.info(f"[EXPORT] {len(all_mcq)} MCQ answers loaded: {(t4-t3)*1000:.0f}ms")

    # 5. All written answers in one query
    stmt = select(WrittenAnswer).where(WrittenAnswer.result_id.in_(result_ids)).order_by(WrittenAnswer.question_number)
    res = await db.execute(stmt)
    all_written = res.scalars().all()
    written_by_result = {}
    for ans in all_written:
        written_by_result.setdefault(ans.result_id, []).append(ans)
    t5 = time.time()
    logger.info(f"[EXPORT] {len(all_written)} written answers loaded: {(t5-t4)*1000:.0f}ms")

    # 6. Answer key
    stmt = select(AnswerKey).where(AnswerKey.test_id == test_id)
    res = await db.execute(stmt)
    answer_key = res.scalars().first()
    written_answers_key = (answer_key.written_questions or {}) if answer_key else {}

    total = time.time() - t0
    logger.info(f"[EXPORT] Total DB load: {total*1000:.0f}ms for {len(results)} results (6 queries)")

    return test, results, users_dict, mcq_by_result, written_by_result, written_answers_key


def _build_row_data(user, result_record, mcq_answers, written_answers, written_answers_key):
    """Build a single row of data for export (shared between Excel and PDF)."""
    student_info = f"{user.full_name or ''} {user.surname or ''} - {user.region or ''}"
    
    # MCQ answers as 0/1 (Q1-Q35)
    mcq_dict = {ans.question_number: ans for ans in mcq_answers}
    mcq_values = []
    for q_num in range(1, 36):
        ans = mcq_dict.get(q_num)
        mcq_values.append(1 if (ans and ans.is_correct) else 0)
    
    # Written answers as 0/1 per sub-part (Q36a, Q36b, ... Q45a, Q45b)
    written_dict = {ans.question_number: ans for ans in written_answers}
    written_values = []
    for q_num in range(36, 46):
        written_ans = written_dict.get(q_num)
        correct_ans = written_answers_key.get(str(q_num), {})
        
        if written_ans and written_ans.student_answer:
            try:
                student = _json.loads(written_ans.student_answer)
            except (ValueError, TypeError):
                student = {}
            
            s_a = _norm(str(student.get('a', '')))
            c_a = _norm(str(correct_ans.get('a', '')))
            written_values.append(1 if _match(s_a, c_a) else 0)
            
            s_b = _norm(str(student.get('b', '')))
            c_b = _norm(str(correct_ans.get('b', '')))
            written_values.append(1 if _match(s_b, c_b) else 0)
        else:
            written_values.append(0)  # Q_a
            written_values.append(0)  # Q_b
    
    return student_info, result_record.total_score, mcq_values, written_values


async def export_results_to_excel(db: AsyncSession, test_id: UUID, filepath: str) -> str:
    """
    Export test results to Excel file.
    Shows 0 (wrong) / 1 (correct) for each question.
    """
    t_start = time.time()
    
    # Batch load all data
    test, results, users_dict, mcq_by_result, written_by_result, written_answers_key = \
        await _load_export_data(db, test_id)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    
    # Headers
    headers = ["Talaba", "Correct"]
    for i in range(1, 36):
        headers.append(f"Q{i}")
    for i in range(36, 46):
        headers.extend([f"Q{i}a", f"Q{i}b"])
    
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows - no more per-row DB queries!
    t_rows = time.time()
    for result_record in results:
        user = users_dict.get(result_record.user_id)
        if not user:
            continue
        
        mcq_answers = mcq_by_result.get(result_record.id, [])
        written_answers = written_by_result.get(result_record.id, [])
        
        student_info, total_score, mcq_values, written_values = \
            _build_row_data(user, result_record, mcq_answers, written_answers, written_answers_key)
        
        row_data = [student_info, total_score] + mcq_values + written_values
        ws.append(row_data)
    
    t_excel = time.time()
    logger.info(f"[EXPORT] Excel rows built: {(t_excel-t_rows)*1000:.0f}ms")
    
    # Auto-size columns
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 10
    
    # Wider column for Talaba
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 10
    
    # Save
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    
    total = time.time() - t_start
    logger.info(f"[EXPORT] Excel export TOTAL: {total*1000:.0f}ms for {len(results)} results")
    
    return filepath


async def export_results_to_pdf(db: AsyncSession, test_id: UUID, filepath: str) -> str:
    """
    Export test results to PDF file.
    Shows 0 (wrong) / 1 (correct) for each question, matching Excel format.
    """
    from reportlab.lib.pagesizes import A3
    
    t_start = time.time()
    
    # Batch load all data
    test, results, users_dict, mcq_by_result, written_by_result, written_answers_key = \
        await _load_export_data(db, test_id)
    
    # Create PDF - use landscape A3 to fit all columns
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    doc = SimpleDocTemplate(
        filepath,
        pagesize=landscape(A3),
        leftMargin=0.3*inch,
        rightMargin=0.3*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1a365d'),
        spaceAfter=20,
        alignment=1  # Center
    )
    
    # Title
    title = Paragraph(f"Test Results: {test.title}", title_style)
    elements.append(title)
    
    # Summary line
    summary_style = ParagraphStyle(
        'Summary',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        alignment=1
    )
    summary = Paragraph(
        f"Students: {len(results)} | Code: {test.test_code} | Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        summary_style
    )
    elements.append(summary)
    elements.append(Spacer(1, 0.3 * inch))
    
    # Headers
    headers = ["Talaba", "Bal"]
    for i in range(1, 36):
        headers.append(f"Q{i}")
    for i in range(36, 46):
        headers.extend([f"{i}a", f"{i}b"])
    
    table_data = [headers]
    
    # Data rows - no more per-row DB queries!
    t_rows = time.time()
    for result_record in results:
        user = users_dict.get(result_record.user_id)
        if not user:
            continue
        
        mcq_answers = mcq_by_result.get(result_record.id, [])
        written_answers = written_by_result.get(result_record.id, [])
        
        student_info, total_score, mcq_values, written_values = \
            _build_row_data(user, result_record, mcq_answers, written_answers, written_answers_key)
        
        row = [student_info, str(total_score)]
        row.extend(str(v) for v in mcq_values)
        row.extend(str(v) for v in written_values)
        
        table_data.append(row)
    
    t_pdf_rows = time.time()
    logger.info(f"[EXPORT] PDF rows built: {(t_pdf_rows-t_rows)*1000:.0f}ms")
    
    # Calculate column widths
    talaba_width = 2.2 * inch
    q_width = 0.28 * inch
    bal_width = 0.35 * inch
    
    col_widths = [talaba_width, bal_width]
    col_widths.extend([q_width] * 55)
    
    results_table = Table(table_data, colWidths=col_widths)
    
    # Base table style
    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 6),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        # Left-align Talaba column
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        # Alternate row shading for readability
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
    ]
    
    results_table.setStyle(TableStyle(style_commands))
    elements.append(results_table)
    
    t_style = time.time()
    logger.info(f"[EXPORT] PDF styling: {(t_style-t_pdf_rows)*1000:.0f}ms")
    
    # Build PDF
    doc.build(elements)
    
    total = time.time() - t_start
    logger.info(f"[EXPORT] PDF export TOTAL: {total*1000:.0f}ms for {len(results)} results")
    
    return filepath
